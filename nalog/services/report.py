import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .schemas import InputRecord


class Report:
    def __init__(self, media_root, filename):
        self.__input_file_path = f"{media_root}/{filename}"
        self.__report_file_path = f"{media_root}//report.xlsx"

    def save_report(self):
        dict_rows = self.__load_data()
        self.__save_data(dict_rows)
        os.remove(self.__input_file_path)

    def __load_data(self) -> list[dict]:
        """
        Загрузка данных с файла
        """
        load_ws = load_workbook(self.__input_file_path).active
        result = []
        for row in load_ws.iter_rows(min_row=3, max_col=6, values_only=True):
            try:
                input_dict = InputRecord(
                    branch=row[0],
                    employee=row[1],
                    tax_base=row[4],
                    actual_tax_calculated=row[5],
                    correct_tax_calculated=0.0,
                    difference=0.0,
                )
                input_dict.calculated_empty_fields()
                result.append(input_dict.model_dump())
            except ValueError as e:
                print(e)
        return result

    def __save_data(self, data: list[dict]) -> None:
        """
        Формирование отчёта с сортировкой
        """
        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = "Отчёт"
        out_ws = self.__creat_sheet_header(out_ws)
        sorted_data = sorted(data, key=lambda d: d["F"], reverse=True)
        for item in sorted_data:
            out_ws.append(item)
        out_wb.save(self.__report_file_path)

    def __creat_sheet_header(self, ws: Worksheet) -> Worksheet:
        """
        Создание шапки отчёта
        """
        ws.merge_cells("A1:A2")
        megre_cell = ws["A1"]
        megre_cell.value = "Филиал"
        self.__creat_style(megre_cell)

        ws.merge_cells("B1:B2")
        megre_cell = ws["B1"]
        megre_cell.value = "Сотрудник"
        self.__creat_style(megre_cell)

        ws.merge_cells("C1:C2")
        megre_cell = ws["C1"]
        megre_cell.value = "Налоговая база"
        self.__creat_style(megre_cell)

        ws.merge_cells("F1:F2")
        megre_cell = ws["F1"]
        megre_cell.value = "Отклонения"
        self.__creat_style(megre_cell)

        ws.merge_cells("D1:E1")
        megre_cell = ws["D1"]
        megre_cell.value = "Налог"
        self.__creat_style(megre_cell)

        cell = ws["D2"]
        cell.value = "Исчислено всего"
        self.__creat_style(cell)

        cell = ws["E2"]
        cell.value = "Исчислено всего по формуле"
        self.__creat_style(cell)

        ws.row_dimensions[2].height = 30
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 15
        return ws

    @staticmethod
    def __creat_style(cell: tuple) -> None:
        """
        Задание стиля для ячейки
        """
        cell.fill = PatternFill("solid", fgColor="CBE4E5")
        cell.font = Font(bold=True, color="000000", name="Arial", size=10)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrapText=True
        )
