import os
from datetime import datetime

from django.conf import settings
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from ..exeptions import NotXLSXFile
from .config import service_settings
from .schemas import Row


class Report:
    __report_file_path: str

    def __init__(self, file_storage: FileSystemStorage):
        self.__file_storage = file_storage

    def save_file(self, file) -> str:
        if file.name.endswith(".xlsx"):
            return self.__file_storage.save(file.name, file)
        else:
            raise NotXLSXFile

    def __delete_file(self, file_name: str) -> None:
        self.__file_storage.delete(file_name)

    def get_report_file_path(self) -> str:
        return self.__report_file_path

    @staticmethod
    def __file_path(file_name: str) -> str:
        now = datetime.now().strftime(service_settings.report.prefix_format)
        name = now + file_name
        path = os.path.join(settings.MEDIA_ROOT, name)
        return path

    def generate_report(self, file_name: str):
        try:
            wb = load_workbook(self.__file_storage.path(file_name))
            data_sheet = wb.active
            row_schemas = self.__validate_rows(data_sheet)
            sorted_rows_dict = self.__sort_rows(row_schemas)
            report_sheet = wb.create_sheet(service_settings.report.sheet.name)
            report_sheet = self.__fill_report_sheet(
                report_sheet, sorted_rows_dict
            )
            self.__coloring_deviations(report_sheet)
            self.__report_file_path = self.__file_path(file_name)
            wb.save(self.__report_file_path)
        except Exception as e:
            print(e)
        finally:
            self.__delete_file(file_name)

    @staticmethod
    def __validate_rows(ws: Worksheet) -> list[Row]:
        """
        Загрузка и валидация данных из файла
        """
        tmp = []
        for row in ws.iter_rows(
            **service_settings.file.iter_config, values_only=True
        ):
            try:
                schema = Row(
                    branch=row[service_settings.file.column["branch"]],
                    employee=row[service_settings.file.column["employee"]],
                    tax_base=row[service_settings.file.column["tax_base"]],
                    actual_tax_calculated=row[
                        service_settings.file.column["actual_tax_calculated"]
                    ],
                )
                tmp.append(schema.calculated_empty_fields())
            except ValueError as e:
                print(e)
        return tmp

    @staticmethod
    def __sort_rows(row_schemas: list[Row]) -> list[dict]:
        """
        Сортировка отчёта по полю "Отклонения"
        """
        tmp = []
        for schema in row_schemas:
            tmp.append(schema.model_dump())
        return sorted(tmp, key=lambda d: d["F"], reverse=True)

    def __fill_report_sheet(
        self, ws: Worksheet, rows_dict: list[dict]
    ) -> Worksheet:
        """
        Заполнение ячеек отчёта
        """
        ws = self.__create_sheet_header(ws)
        for item in rows_dict:
            ws.append(item)
        return ws

    @staticmethod
    def __coloring_deviations(ws: Worksheet) -> None:
        """
        Заливка ячеек цветом
        Корректный расчёт -> зелёный
        Некорректный расчёт -> красный
        """
        for row in ws.iter_rows(min_row=3, min_col=6, max_col=6):
            cell = row[0]
            if cell.value != 0:
                cell.fill = PatternFill(
                    **service_settings.report.sheet.pattern_fill.incorrect
                )
            else:
                cell.fill = PatternFill(
                    **service_settings.report.sheet.pattern_fill.correct
                )

    def __create_sheet_header(self, ws: Worksheet) -> Worksheet:
        """
        Создание шапки отчёта
        """
        for k, v in service_settings.report.sheet.cells.merge.items():
            ws.merge_cells(k)
            first_cell, _ = k.split(":")
            self.__set_sheet_header_cell_style(ws, first_cell, v)
        for k, v in service_settings.report.sheet.cells.single.items():
            self.__set_sheet_header_cell_style(ws, k, v)
        for k, v in service_settings.report.sheet.cells.width.items():
            ws.column_dimensions[k].width = v
        for k, v in service_settings.report.sheet.cells.height.items():
            ws.row_dimensions[k].height = v
        return ws

    @staticmethod
    def __set_sheet_header_cell_style(
        ws: Worksheet, cell_key: str, cell_value: str
    ) -> None:
        """
        Задание стиля ячейки для шапки таблицы
        """
        cell = ws[cell_key]
        cell.value = cell_value
        cell.fill = PatternFill(
            **service_settings.report.sheet.pattern_fill.header
        )
        cell.font = Font(**service_settings.report.sheet.font.header)
        cell.alignment = Alignment(
            **service_settings.report.sheet.alignment.header
        )
